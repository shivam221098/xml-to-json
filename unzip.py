import json
import openpyxl
from time import sleep
from zipfile import ZipFile

from xmltodict import parse
flag = False


def major_descriptor(yn):
    if yn.lower() == 'y':
        return True
    return False


"""with ZipFile('test.zip', 'r') as zip_file:
    with zip_file.open('pubmed21n0020.xml') as file:
        with open('pubmed2.json', 'w', encoding='utf-8') as json_file:
            json_file.write(json.dumps(parse(file)))"""

with open('pubmed2.json', 'r', encoding='utf-8') as file:
    data = json.load(file)
print("Temporary file created!")
wb = openpyxl.Workbook()

ws = wb.active

ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], \
ws['F1'] = 'pmid', "descriptor_uid", 'major_descriptor', "article_title", \
           "date_created", "author_ordinality"
ws['G1'] = 'affiliation'
ws['H1'] = 'affiliation_ordinality'
ws['I1'] = 'initials'
ws['J1'] = 'fore_name'
ws["K1"] = 'last_name'
ws['L1'] = 'date_revised'
ws['M1'] = 'issn'
ws['N1'] = 'issn_type'
ws['O1'] = 'cited_medium'
ws["P1"] = 'volume'
ws['Q1'] = 'issue'
ws['R1'] = 'year'
ws['S1'] = 'month'
ws['T1'] = 'title'
ws['U1'] = 'iso_abbreviation'
ws['V1'] = 'nlm_uid'
ws['W1'] = 'publication_type'
ws['X1'] = 'publication_type_ui'
ws['Y1'] = 'publication_type_ordinality'

count = 1
print("Process started!")
data = data.get("PubmedArticleSet").get('PubmedArticle')
for value in data:

    PMID = value.get("MedlineCitation").get("PMID").get("#text")
    DescriptorUIDList = value.get("MedlineCitation").get("MeshHeadingList").get("MeshHeading")
    ArticleTitle = value.get("MedlineCitation").get("Article").get('ArticleTitle')
    DateCreated = value.get("MedlineCitation").get("DateCompleted").get("Day") + '/' + \
                  value.get("MedlineCitation").get("DateCompleted").get("Month") + '/' + \
                  value.get("MedlineCitation").get("DateCompleted").get("Year")
    AuthorList = value.get("MedlineCitation").get("Article").get("AuthorList")

    if AuthorList is None:
        Author = [{"Initials": "######", "ForeName": "######", "LastName": "#######"}]
    else:
        Author = AuthorList.get("Author")

    DateRevised = value.get("MedlineCitation").get("DateRevised").get("Day") + '/' + \
                  value.get("MedlineCitation").get("DateRevised").get("Month") + '/' + \
                  value.get("MedlineCitation").get("DateRevised").get("Year")

    ISSN = value.get("MedlineCitation").get("Article").get('Journal').get("ISSN")

    #ISSN = None
    if ISSN is not None:
        issn = ISSN.get("#text")
        issn_type = ISSN.get("@IssnType")
    else:
        issn = "######"
        issn_type = "######"

    CitedMedium = value.get("MedlineCitation").get("Article").get("Journal").get("JournalIssue").get("@CitedMedium")
    Volume = value.get("MedlineCitation").get("Article").get("Journal").get("JournalIssue").get("Volume")
    Issue = value.get("MedlineCitation").get("Article").get("Journal").get("JournalIssue").get("Issue")
    year_month = value.get("MedlineCitation").get("Article").get("Journal").get("JournalIssue").get("PubDate").get("MedlineDate")

    if year_month is None:
        Year, Month = "####", "##"
    else:
        if len(year_month.split()) == 2:
            Year, Month = list(year_month.split())
        else:
            Year, Month = "####", "##"

    Title = value.get("MedlineCitation").get("Article").get('Journal').get("Title")
    ISOAbbreviation = value.get("MedlineCitation").get("Article").get('Journal').get("ISOAbbreviation")
    NlmUniqueId = value.get("MedlineCitation").get("MedlineJournalInfo").get("NlmUniqueID")
    PublicationTypeList = value.get("MedlineCitation").get("Article").get("PublicationTypeList").get("PublicationType")
    if type(PublicationTypeList) is not list:
        PublicationTypeList = [PublicationTypeList]

    if type(Author) is not list:
        Author = [Author]

    if type(DescriptorUIDList) is not list:
        DescriptorUIDList = [DescriptorUIDList]

    """print(PubmedArticle)
    print(DescriptorUIDList)
    print(ArticleTitle)
    print(DateCreated)
    print(Author)
    print(Year, Month)
    print(Title)
    print(ISOAbbreviation)
    print(Issue)
    print(NlmUniqueId)
    print(PublicationTypeList)"""
    count_1 = count_2 = 1  # Count_1 = author_ordinality, count_2 = publication_type_ordinality
    for description in DescriptorUIDList:
        count_1 = 1
        for author in Author:
            count_2 = 1
            for publication in PublicationTypeList:
                major = major_descriptor(description.get("DescriptorName").get("@MajorTopicYN"))
                ws.append([PMID, description.get("DescriptorName").get("@UI"),
                           major, ArticleTitle, DateCreated, count_1, "######", "######",
                           author.get("Initials"), author.get("ForeName"), author.get("LastName"), DateRevised,
                           issn, issn_type, CitedMedium, Volume, Issue, Year, Month, Title, ISOAbbreviation,
                           NlmUniqueId, publication.get("#text"), publication.get("@UI"), count_2])
                count_2 += 1
            count_1 += 1
        print("Count:", count)
        count += 1
        if count == 10000:
            flag = True
    if flag:
        break

wb.save('sample.csv')
print("Process completed!")
